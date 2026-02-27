import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, date

# Load the workbook
wb = openpyxl.load_workbook('Gantt_chart_AUTO.xlsx')
sheet = wb.active

# Current date
today = date(2026, 2, 20)

# Start date of Gantt chart (assuming Monday 2026-01-06)
gantt_start_date = date(2026, 1, 6)

# Column indices (0-based)
# A=0: WBS, B=1: Title, C=2: Owner, D=3: Start, E=4: End, F=5: Duration, G=6: PCT
# H=7: Phase1, I=8: Phase2, J=9: Phase3, K=10: Phase4
# L=11: Week1 Mon, M=12: Tue, N=13: Wed, O=14: Thu, P=15: Fri
# Q=16: Week2 Mon, etc.

# Debug: print row 11 and 12
for r in [11, 12]:
    print(f"Row {r} data:")
    for col in range(1, 20):  # First 20 columns
        cell = sheet.cell(row=r, column=col)
        print(f"Col {col} ({chr(64+col)}): {cell.value}")
    print()

# Find task rows (WBS in column 2)
task_rows = []
for row in range(10, sheet.max_row + 1):
    wbs_cell = sheet.cell(row=row, column=2)  # Column B
    val = wbs_cell.value
    if val is not None:
        val_str = str(val)
        if '.' in val_str and val_str.replace('.', '').replace('-', '').isdigit():
            task_rows.append(row)

print(f"Found {len(task_rows)} tasks: {task_rows}")

# Colors
colors = {
    'completed': '90EE90',   # light green
    'done': 'ADD8E6',        # light blue
    'remaining': 'D3D3D3'    # light grey
}

# Process each task
for row in task_rows:
    start_cell = sheet.cell(row=row, column=5)  # E
    end_cell = sheet.cell(row=row, column=6)    # F
    pct_cell = sheet.cell(row=row, column=8)    # H

    start_val = start_cell.value
    end_val = end_cell.value
    pct_complete = pct_cell.value if pct_cell.value else 0

    # Parse dates
    try:
        if isinstance(start_val, str):
            start_date = datetime.strptime(start_val, '%Y-%m-%d %H:%M:%S').date()
        elif isinstance(start_val, datetime):
            start_date = start_val.date()
        else:
            continue

        if isinstance(end_val, str):
            end_date = datetime.strptime(end_val, '%Y-%m-%d %H:%M:%S').date()
        elif isinstance(end_val, datetime):
            end_date = end_val.date()
        else:
            continue
    except Exception as e:
        print(f"Error parsing dates for row {row}: {e}")
        continue

    # Determine coloring
    if pct_complete >= 1:
        # Completed: light green for entire bar
        bar_color = 'completed'
    else:
        # Partial: will color cells individually
        bar_color = None  # special handling

    # Color the Gantt cells for this task
    from datetime import timedelta
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() < 5:  # Weekday only
            # Calculate column
            days_diff = (current_date - gantt_start_date).days
            weeks_from_start = days_diff // 7
            day_in_week = current_date.weekday()
            col = 12 + weeks_from_start * 5 + day_in_week

            if col <= sheet.max_column:
                cell = sheet.cell(row=row, column=col)
                if bar_color:
                    cell.fill = PatternFill(start_color=colors[bar_color], end_color=colors[bar_color], fill_type='solid')
                else:
                    # Partial: determine color based on progress
                    total_days = (end_date - start_date).days + 1
                    progress_days = int(total_days * pct_complete)
                    progress_date = start_date + timedelta(days=progress_days - 1) if progress_days > 0 else start_date - timedelta(days=1)
                    
                    if current_date <= progress_date:
                        color = colors['done']  # light blue
                    else:
                        color = colors['remaining']  # light grey
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        current_date += timedelta(days=1)

# Save the workbook
wb.save('Gantt_chart_AUTO_colored_v2.xlsx')
print("Gantt chart colored and saved as 'Gantt_chart_AUTO_colored_v2.xlsx'")